import openpyxl as op
from time import sleep
import requests
import cloudscraper
import re
from bs4 import BeautifulSoup
import os
from selenium import webdriver

from requests_html import HTMLSession
from requests_html import AsyncHTMLSession
import shutil
import time
import random

def Soup_Extractor(url):
    scraper = cloudscraper.create_scraper(delay=10, browser={'custom': 'ScraperBot/1.0', })
    response = scraper.get(url)
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}

    session = requests.Session()
    response = session.get(url, headers=headers)

    content = response.content



    soup = BeautifulSoup(content, 'html.parser')

    return(soup)








folder = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads')
file_name = os.path.join(folder, 'Upwork_task_ships.xlsx')
wb = op.load_workbook(file_name)
ws = wb['Feuil1']

max_row = ws.max_row
max_column = ws.max_column


i = 2
url = 'http://www.shipbucket.com/drawings?category=real'

url_0 = "http://www.shipbucket.com"
while i < max_row +1:


    country = ws.cell(row = i, column = 1).value

    soup = Soup_Extractor(url)
    for tag in soup.find_all('a'):
        if tag.text.strip() == country:
            url1 = url_0 + tag['href']
            print(url1, country )



    for j in range(i , max_row+1):
        type  = ws.cell(row = j, column = 3).value
        type = type.split(" ")


        soup = Soup_Extractor(url1)
        for tag in soup.find_all('a'):
            if tag.text.strip() in type:
                url2 = url_0 + tag['href']
                print(url2, type)




        for k in range(j, max_row + 1):
            print((i + j + k) % 20)

            if (i + j + k) % 20 == 0:
                sleep(random.randrange(50, 150))


            name = ws.cell(row=k, column=2).value
            name = name.split(" ")

            soup = Soup_Extractor(url2)

            n = 5
            for tag in soup.find_all('a'):
                text_0 = tag.text.strip()
                texts = text_0.split(" ")
                text_0 = text_0 + ".png"

                for text in texts:
                    if text in name:            #Checks for element with required name to provide url link and image link
                        url3 = url_0 + tag['href']
                        file_link = url_0 + tag.find('img')['src']


                        # Downloading images from the site
                        path = os.path.join(folder, text_0)
                        r = requests.get(file_link, stream=True)
                        if r.status_code == 200:
                            with open(path, 'wb') as f:
                                r.raw.decode_content = True
                                shutil.copyfileobj(r.raw, f)



                        #Writing data to excel file
                        ws.cell(row = k, column = n).value = text_0
                        ws.cell(row = k, column = n + 1).value = url3
                        n = n + 2





                        print(url3, name, texts)
                        print(file_link)
                        break



            print(country, type, name)
            wb.save(file_name)










            if type != ws.cell(row=j + 1, column=1).value:
                i=j + 1
                break



        if country != ws.cell(row = i+1, column = 1).value:
            break

wb.save(file_name)